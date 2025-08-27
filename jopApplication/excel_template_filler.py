"""
입사지원서 자동 작성 도구 (플레이스홀더 방식)
Raw 데이터 엑셀 파일의 각 행을 읽어서 템플릿에 자동으로 매핑하여 개별 지원서 생성

사용법:
1. templates/ 폴더에 플레이스홀더가 포함된 템플릿 파일 배치
2. raw_data/ 폴더에 원천 데이터 엑셀 파일 배치  
3. config.json에서 매핑 설정
4. python excel_template_filler.py 실행

플레이스홀더 문법:
- 기본: {{필드명}}
- 변환: {{필드명|변환1|변환2:인자}}

변환 종류:
- trim: 앞뒤 공백 제거
- upper/lower: 대소문자 변환
- digits: 숫자만 추출 (전화번호용)
- zfill:N: N자리까지 0으로 채움
- date:입력포맷->출력포맷: 날짜 형식 변환
- map:키=값,키=값: 값 매핑 (성별 등)
- default:기본값: 빈 값일 때 기본값 사용
"""

import os
import re
import json
import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from pathlib import Path
import shutil
from copy import deepcopy
import xlwings as xw

class ExcelTemplateFiller:
    def __init__(self, config_path="config.json"):
        """초기화"""
        self.config = self.load_config(config_path)
        self.placeholder_pattern = re.compile(r"\{\{\s*([^}|]+)\s*(\|[^}]*)?\}\}")
        
    def load_config(self, config_path):
        """설정 파일 로드"""
        if not os.path.exists(config_path):
            # 기본 설정 파일 생성
            default_config = {
                "template_file": "../templates/application_template.xlsx",
                "raw_data_file": "../raw_data/applicants.xlsx", 
                "raw_data_sheet": "공고별 지원자 관리",
                "output_dir": "../output",
                "filename_pattern": "{이름}_입사지원서.xlsx",
                "save_pdf": True,  # PDF 저장 옵션 추가
                "encoding": "utf-8",
                # 지원자 사진 관련 설정
                "images_dir": "../images",  # 지원자 사진 폴더
                "photo_field": "수험번호",  # 사진 파일명의 기준 필드
                "photo_extensions": [".png", ".jpg", ".jpeg"],  # 지원 이미지 형식
                "photo_placeholder": "{{사진}}",  # 템플릿에서 사진 위치 지정
                "photo_width": 121,   # 사진 너비 (픽셀) - 열너비 17.25 * 7
                "photo_height": 156   # 사진 높이 (픽셀) - 행높이와 동일
            }
            with open(config_path, 'w', encoding='utf-8') as f:
                json.dump(default_config, f, ensure_ascii=False, indent=2)
            print(f"기본 설정 파일을 생성했습니다: {config_path}")
            return default_config
        
        with open(config_path, 'r', encoding='utf-8') as f:
            config = json.load(f)
            
        # 새로운 옵션이 없으면 기본값 추가
        if "save_pdf" not in config:
            config["save_pdf"] = True
            
        # 이미지 관련 기본값 추가
        if "images_dir" not in config:
            config["images_dir"] = "../images"
        if "photo_field" not in config:
            config["photo_field"] = "수험번호"
        if "photo_extensions" not in config:
            config["photo_extensions"] = [".png", ".jpg", ".jpeg"]
        if "photo_placeholder" not in config:
            config["photo_placeholder"] = "{{사진}}"
        if "photo_width" not in config:
            config["photo_width"] = 121
        if "photo_height" not in config:
            config["photo_height"] = 156
            
        return config
    
    def apply_transforms(self, value, pipe_spec, context=None):
        """파이프라인 변환 적용"""
        if value is None or pd.isna(value):
            value = ""
        
        s = str(value).strip()
        
        if not pipe_spec:
            return s
            
        # 파이프 구분하여 변환 단계별 적용
        steps = [p.strip() for p in pipe_spec.strip("|").split("|") if p.strip()]
        
        for step in steps:
            if step == "trim":
                s = s.strip()
            elif step == "upper":
                s = s.upper()
            elif step == "lower":
                s = s.lower()
            elif step.startswith("zfill:"):
                try:
                    n = int(step.split(":")[1])
                    s = s.zfill(n)
                except:
                    pass
            elif step == "digits":
                s = re.sub(r"\D+", "", s)
            elif step.startswith("date:"):
                # date:%Y-%m-%d->%Y.%m.%d
                try:
                    spec = step.split(":", 1)[1]
                    if "->" in spec:
                        src_fmt, dst_fmt = spec.split("->")
                        dt = datetime.datetime.strptime(s, src_fmt)
                        s = dt.strftime(dst_fmt)
                except Exception as e:
                    print(f"날짜 변환 실패: {s} -> {step}, 오류: {e}")
                    pass
            elif step.startswith("map:"):
                # map:남=Male,여=Female
                try:
                    pairs = step.split(":", 1)[1]
                    mapping = {}
                    for pair in pairs.split(","):
                        if "=" in pair:
                            key, val = pair.split("=", 1)
                            mapping[key.strip()] = val.strip()
                    s = mapping.get(s, s)
                except Exception as e:
                    print(f"매핑 변환 실패: {s} -> {step}, 오류: {e}")
                    pass
            elif step.startswith("default:"):
                # 값이 비어있으면 기본값 사용
                try:
                    default_val = step.split(":", 1)[1]
                    if s.strip() == "":
                        s = default_val
                except:
                    pass
            elif step.startswith("prefix:"):
                try:
                    prefix = step.split(":", 1)[1]
                    s = prefix + s
                except:
                    pass
            elif step.startswith("suffix:"):
                try:
                    suffix = step.split(":", 1)[1]
                    s = s + suffix
                except:
                    pass
            elif step == "extract_age":
                # "만 31세(32)" -> "만 31세(32)"
                try:
                    import re
                    # 전체 패턴 매칭하여 닫는 괄호까지 포함
                    match = re.search(r'만 \d+세\(\d+\)', s)
                    if match:
                        s = match.group(0)
                    else:
                        # 닫는 괄호가 없는 경우 추가
                        match = re.search(r'만 \d+세\(\d+', s)
                        if match:
                            s = match.group(0) + ")"
                        else:
                            # 기본 패턴으로 시도
                            match = re.search(r'(만 \d+세)', s)
                            if match:
                                s = match.group(1)
                except:
                    pass
            elif step.startswith("split_line:"):
                # split_line:0 (첫 번째 줄), split_line:1 (두 번째 줄), split_line:2 (세 번째 줄)
                try:
                    line_index = int(step.split(":")[1])
                    # 줄바꿈으로 분리 (Windows \r\n, Unix \n 모두 처리)
                    lines = s.replace('\r\n', '\n').split('\n')
                    if line_index < len(lines):
                        s = lines[line_index].strip()
                        print(f"  split_line:{line_index} -> '{s}'")  # 디버그
                    else:
                        s = ""  # 해당 줄이 없으면 빈 값
                        print(f"  split_line:{line_index} -> 빈 값 (줄 없음)")  # 디버그
                except Exception as e:
                    print(f"split_line 변환 실패: {s} -> {step}, 오류: {e}")
                    pass
            elif step.startswith("combine:"):
                # combine:복무종료일,~,%Y-%m-%d->%y.%m.%d
                try:
                    spec = step.split(":", 1)[1]
                    parts = spec.split(",")
                    if len(parts) >= 2:
                        other_field = parts[0].strip()
                        separator = parts[1].strip()
                        
                        # 다른 필드 값 가져오기
                        other_value = str(context.get(other_field, "")).strip()
                        
                        # 세 번째 파라미터가 있으면 변환 또는 날짜 포맷으로 처리
                        if len(parts) >= 3:
                            third_param = parts[2].strip()
                            
                            # 날짜 포맷인지 확인 (-> 포함 여부)
                            if "->" in third_param:
                                # 날짜 포맷 변환
                                src_fmt, dst_fmt = third_param.split("->")
                                try:
                                    # 시작일 변환
                                    if s:
                                        start_dt = datetime.datetime.strptime(s, src_fmt)
                                        s = start_dt.strftime(dst_fmt)
                                    # 종료일 변환  
                                    if other_value:
                                        end_dt = datetime.datetime.strptime(other_value, src_fmt)
                                        other_value = end_dt.strftime(dst_fmt)
                                except Exception as e:
                                    print(f"combine 날짜 변환 실패: {s}, {other_value} -> {third_param}, 오류: {e}")
                            else:
                                # 변환 이름인 경우 (extract_age 등)
                                other_value = self.apply_transforms(other_value, "|" + third_param, context)
                        
                        # 결합
                        if s and other_value:
                            s = f"{s}{separator}{other_value}"
                            print(f"  combine 결합: '{s}' + '{separator}' + '{other_value}' = '{s}'")  # 디버그
                        elif s:
                            s = s  # 시작일만 있는 경우
                        elif other_value:
                            s = other_value  # 종료일만 있는 경우
                        else:
                            s = ""  # 둘 다 없는 경우
                
                except Exception as e:
                    print(f"combine 변환 실패: {s} -> {step}, 오류: {e}")
                    pass
        
        return s
    
    def replace_placeholders_in_cell(self, cell, context):
        """셀의 플레이스홀더를 실제 값으로 치환"""
        if not isinstance(cell.value, str):
            return
            
        original = cell.value
        
        def replace_func(match):
            field = match.group(1).strip()
            pipe = match.group(2) or ""
            value = context.get(field, "")
            return self.apply_transforms(value, pipe, context)
        
        new_value = self.placeholder_pattern.sub(replace_func, original)
        
        if new_value != original:
            cell.value = new_value
            print(f"  셀 치환: '{original}' -> '{new_value}'")
    
    def replace_placeholders_in_string(self, text, context):
        """문자열의 플레이스홀더를 실제 값으로 치환 (xlwings용)"""
        if not isinstance(text, str):
            return text
            
        def replace_func(match):
            field = match.group(1).strip()
            pipe = match.group(2) or ""
            value = context.get(field, "")
            return self.apply_transforms(value, pipe, context)
        
        return self.placeholder_pattern.sub(replace_func, text)
    
    def find_applicant_photo(self, context):
        """지원자 사진 파일 찾기"""
        images_dir = self.config["images_dir"]
        photo_field = self.config["photo_field"]
        extensions = self.config["photo_extensions"]
        
        # 수험번호 추출
        if photo_field not in context:
            return None
            
        exam_number = str(context[photo_field]).strip()
        if not exam_number:
            return None
        
        # images 디렉토리에서 해당 수험번호로 시작하는 파일 찾기
        images_path = Path(images_dir)
        if not images_path.exists():
            print(f"  이미지 폴더가 없습니다: {images_dir}")
            return None
        
        # 패턴: 수험번호_*.png/jpg/jpeg
        for ext in extensions:
            pattern = f"{exam_number}_*{ext}"
            matching_files = list(images_path.glob(pattern))
            
            if matching_files:
                photo_path = matching_files[0]  # 첫 번째 매칭 파일 사용
                print(f"  지원자 사진 발견: {photo_path.name}")
                return str(photo_path.absolute())
        
        print(f"  지원자 사진 없음: {exam_number}_*")
        return None

    def insert_photo_xlwings(self, sheet, photo_path, target_cell):
        """xlwings를 사용하여 지원자 사진을 병합된 셀 범위에 맞춰 삽입 (개선된 버전)"""
        try:
            if not os.path.exists(photo_path):
                print(f"    사진 파일 없음: {photo_path}")
                return False
            
            # 대상 셀 위치 가져오기
            if "," in target_cell:  # row,col 형식
                row, col = map(int, target_cell.split(","))
                cell_range = sheet.range(row, col)
            else:  # A1 형식
                cell_range = sheet.range(target_cell)
            
            # 사진 틀 영역 설정 (B5:C12 병합된 셀 범위)
            photo_frame_range = sheet.range("B5:C12")
            cell_width = photo_frame_range.width    # 병합된 셀의 실제 너비 (픽셀)
            cell_height = photo_frame_range.height  # 병합된 셀의 실제 높이 (픽셀)
            
            # 사진 삽입 위치는 B5 셀의 위치 사용
            insert_left = photo_frame_range.left
            insert_top = photo_frame_range.top
            
            print(f"    셀 크기: {cell_width:.1f} x {cell_height:.1f} 픽셀")
            
            # 사진 파일의 원본 크기 확인 시도
            try:
                from PIL import Image
                with Image.open(photo_path) as img: # 이미지 파일을 열어서 Image 객체 생성
                    original_width, original_height = img.size
                    print(f"    원본 사진 크기: {original_width} x {original_height} 픽셀")
                    
                    # 비율 유지하면서 셀 크기에 맞춤
                    width_ratio = cell_width / original_width
                    height_ratio = cell_height / original_height
                    scale_ratio = min(width_ratio, height_ratio)  # 작은 비율 선택 (셀을 벗어나지 않도록)
                    
                    final_width = original_width * scale_ratio
                    final_height = original_height * scale_ratio
                    
                    print(f"    조정된 사진 크기: {final_width:.1f} x {final_height:.1f} 픽셀 (비율: {scale_ratio:.3f})")
                    
                    # 중앙 정렬을 위한 오프셋 계산
                    left_offset = (cell_width - final_width) / 2
                    top_offset = (cell_height - final_height) / 2
                    
                    # 이미지 삽입 (병합된 셀 범위에 중앙 정렬)
                    picture = sheet.pictures.add(
                        photo_path,
                        left=insert_left + left_offset,
                        top=insert_top + top_offset,
                        width=final_width,
                        height=final_height
                    )
                    
                    print(f"    사진 삽입 완료: {os.path.basename(photo_path)} -> {target_cell} (중앙 정렬)")
                    return True
                    
            except ImportError:
                print("    PIL(Pillow) 라이브러리가 없어 기본 크기로 삽입합니다.")
            except Exception as pil_error:
                print(f"    사진 크기 분석 실패: {pil_error}")
            
            # PIL이 없거나 실패한 경우 기본 방식 (셀 크기에 맞춤)
            print("    셀 크기에 맞춰 기본 방식으로 삽입...")
            
            # 이미지 삽입 (병합된 셀 크기에 맞춤)
            picture = sheet.pictures.add(
                photo_path,
                left=insert_left,
                top=insert_top,
                width=cell_width,
                height=cell_height
            )
            
            print(f"    사진 삽입 완료 (셀 크기 맞춤): {os.path.basename(photo_path)} -> {target_cell}")
            return True
            
        except Exception as e:
            print(f"    사진 삽입 실패: {e}")
            return False
    
    def save_as_pdf_xlwings(self, excel_path, pdf_path):
        """xlwings를 사용해 Excel을 PDF로 변환 (개선된 버전)"""
        app = None
        wb = None
        try:
            print(f"  PDF 변환 시작: {pdf_path}")
            
            # 절대 경로로 변환
            excel_path_abs = os.path.abspath(excel_path)
            pdf_path_abs = os.path.abspath(pdf_path)
            print(f"  Excel 경로: {excel_path_abs}")
            print(f"  PDF 경로: {pdf_path_abs}")
            
            # Excel 애플리케이션 시작 (백그라운드)
            app = xw.App(visible=False, add_book=False)
            
            # Excel 파일 열기
            wb = app.books.open(excel_path_abs)
            
            # PDF로 저장 - 단순한 방식
            try:
                wb.api.ExportAsFixedFormat(0, pdf_path_abs)
                print(f"  PDF 저장 완료: {pdf_path}")
            except Exception as e1:
                print(f"  전체 워크북 PDF 저장 실패: {e1}")
                # 대안: 첫 번째 시트만 PDF로 저장
                print("  첫 번째 시트만 PDF로 저장 시도...")
                wb.sheets[0].activate()
                wb.sheets[0].api.ExportAsFixedFormat(0, pdf_path_abs)
                print(f"  PDF 저장 완료 (첫 번째 시트만): {pdf_path}")
            
            wb.close()
            
        except Exception as e:
            print(f"  PDF 변환 실패: {e}")
            if wb:
                wb.close()
            raise
        finally:
            if app:
                app.quit()

    def fill_workbook_xlwings(self, template_path, context, output_path):
        """xlwings를 사용한 완벽한 이미지 보존 방식 + PDF 저장"""
        print(f"\n템플릿 처리 (xlwings - 이미지 보존): {template_path}")
        
        # 출력 디렉토리 생성
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        
        # 1단계: 템플릿 파일을 출력 위치로 복사 (이미지 포함)
        try:
            # 파일 속성까지 완전히 복사 (이미지 보존)
            shutil.copy2(template_path, output_path)
            print(f"  템플릿 복사 완료 (이미지 포함): {output_path}")
            
            # 파일 존재 및 크기 확인
            template_size = os.path.getsize(template_path)
            output_size = os.path.getsize(output_path)
            print(f"  원본 크기: {template_size:,} bytes")
            print(f"  복사본 크기: {output_size:,} bytes")
            
            if abs(template_size - output_size) > 1000:  # 1KB 이상 차이나면 경고
                print("  파일 크기 차이 감지 - 이미지 손실 가능성")
            
        except Exception as e:
            print(f"  템플릿 복사 실패: {e}")
            raise
        
        # 2단계: xlwings로 Excel 애플리케이션 제어 (이미지 보존 모드)
        app = None
        wb = None
        try:
            # Excel 애플리케이션 시작 (백그라운드, 이미지 보존 설정)
            app = xw.App(visible=False, add_book=False)
            print("  Excel 애플리케이션 시작 (이미지 보존 모드)")
            
            # 복사된 파일 열기 (읽기/쓰기 모드) - 안전하게 시도
            try:
                wb = app.books.open(output_path, update_links=False)
                print("  파일 로드 완료 (이미지 포함)")
            except Exception as open_error:
                print(f"  update_links=False로 열기 실패, 기본 방식 시도: {open_error}")
                wb = app.books.open(output_path)
                print("  파일 로드 완료 (기본 방식)")
            
            # 이미지 개수 확인 (디버깅용) - 안전하게 시도
            try:
                total_images = 0
                for sheet in wb.sheets:
                    try:
                        sheet_images = len(sheet.pictures)
                        total_images += sheet_images
                        if sheet_images > 0:
                            print(f"  {sheet.name} 시트: {sheet_images}개 이미지 발견")
                    except Exception as sheet_img_error:
                        print(f"  {sheet.name} 시트 이미지 확인 실패: {sheet_img_error}")
                print(f"  총 이미지 개수: {total_images}개")
            except Exception as img_check_error:
                print(f"  이미지 개수 확인 실패: {img_check_error}")
            
            # 지원자 사진 파일 찾기 (한 번만 실행)
            photo_path = self.find_applicant_photo(context)
            
            # 모든 시트에서 플레이스홀더 치환
            for sheet in wb.sheets:
                print(f"  시트 처리: {sheet.name}")
                placeholder_count = 0
                photo_inserted = False
                
                # 사용된 범위 가져오기
                used_range = sheet.used_range
                if used_range is None:
                    continue
                
                # 모든 셀 검사
                for row in range(1, used_range.last_cell.row + 1):
                    for col in range(1, used_range.last_cell.column + 1):
                        cell = sheet.range(row, col)
                        cell_value = cell.value
                        
                        if cell_value and isinstance(cell_value, str) and '{{' in cell_value:
                            original_value = cell_value
                            
                            # {{사진}} 플레이스홀더 처리 (특별 처리)
                            if self.config["photo_placeholder"] in original_value:
                                if photo_path and not photo_inserted:
                                    # 사진 삽입 (셀 위치 직접 사용)
                                    if self.insert_photo_xlwings(sheet, photo_path, f"{row},{col}"):
                                        photo_inserted = True
                                        placeholder_count += 1
                                        print(f"    사진 삽입: {original_value} -> 이미지 파일")
                                    
                                # 플레이스홀더 텍스트 제거
                                cell.value = ""
                            else:
                                # 일반 플레이스홀더 처리
                                new_value = self.replace_placeholders_in_string(original_value, context)
                                
                                if new_value != original_value:
                                    cell.value = new_value
                                    placeholder_count += 1
                                    print(f"    치환 {placeholder_count}: {original_value[:30]}... -> {new_value[:30]}...")
                
                print(f"  {sheet.name} 시트: {placeholder_count}개 플레이스홀더 처리 완료")
            
            # 3단계: Excel 저장 (이미지 보존 확인)
            wb.save()
            print(f"Excel 저장 완료: {output_path}")
            
            # 저장 후 이미지 보존 확인 - 안전하게 시도
            try:
                final_size = os.path.getsize(output_path)
                print(f"  최종 파일 크기: {final_size:,} bytes")
                
                # 이미지 개수 재확인
                total_images_after = 0
                for sheet in wb.sheets:
                    try:
                        total_images_after += len(sheet.pictures)
                    except Exception as sheet_final_error:
                        print(f"  {sheet.name} 시트 최종 이미지 확인 실패: {sheet_final_error}")
                print(f"  저장 후 이미지 개수: {total_images_after}개")
                
            except Exception as final_check_error:
                print(f"  최종 확인 실패: {final_check_error}")
            
            # 4단계: PDF 저장 (설정에 따라)
            save_pdf_option = self.config.get("save_pdf", True)
            print(f"  PDF 저장 설정: {save_pdf_option}")
            
            if save_pdf_option:
                pdf_path = output_path.replace('.xlsx', '.pdf')
                print(f"  PDF 저장 시작: {pdf_path}")
                try:
                    # 절대 경로로 변환 (경로 문제 해결)
                    pdf_path_abs = os.path.abspath(pdf_path)
                    print(f"  절대 경로: {pdf_path_abs}")
                    
                    # PDF로 저장 - 다른 방식 시도
                    wb.api.ExportAsFixedFormat(0, pdf_path_abs)
                    print(f"PDF 저장 완료: {pdf_path}")
                except Exception as e1:
                    print(f"PDF 저장 실패 (방법1): {e1}")
                    # 대안 방법: 각 시트를 개별적으로 저장
                    try:
                        print("  대안 방법 시도: 활성 시트만 PDF로 저장")
                        # 첫 번째 시트를 활성화하고 PDF로 저장
                        wb.sheets[0].activate()
                        wb.sheets[0].api.ExportAsFixedFormat(0, pdf_path_abs)
                        print(f"PDF 저장 완료 (활성 시트만): {pdf_path}")
                    except Exception as e2:
                        print(f"PDF 저장 완전 실패: {e2}")
                        print("해결책: Excel 파일을 수동으로 열어서 '파일 > 내보내기 > PDF 만들기'를 사용해주세요.")
            else:
                print("  PDF 저장 건너뛰기 (설정에서 비활성화)")
            
            wb.close()
            
        except Exception as e:
            print(f"xlwings 처리 실패: {e}")
            if wb:
                wb.close()
            # 복사된 파일 삭제 (실패 시)
            if os.path.exists(output_path):
                os.remove(output_path)
            raise
        finally:
            # Excel 애플리케이션 종료
            if app:
                app.quit()

    def fill_workbook(self, template_path, context, output_path):
        """기본 fill_workbook - xlwings 강제 사용 (이미지 보존)"""
        try:
            # xlwings 방식 우선 시도 (완벽한 이미지 보존)
            print("이미지 보존을 위해 xlwings 사용")
            self.fill_workbook_xlwings(template_path, context, output_path)
        except Exception as e:
            print(f"xlwings 실패: {e}")
            print("이미지 보존이 중요한 경우 Excel이 설치된 환경에서 실행해주세요.")
            
            # 사용자에게 선택권 제공
            print("\n어떻게 처리하시겠습니까?")
            print("1. openpyxl로 대체 (이미지 손실 가능)")
            print("2. 중단 (Excel 환경에서 다시 실행)")
            
            try:
                choice = input("선택 (1 또는 2): ").strip()
                if choice == "2":
                    print("실행 중단. Excel이 설치된 환경에서 다시 실행해주세요.")
                    raise Exception("사용자가 실행을 중단했습니다.")
                elif choice == "1":
                    print("openpyxl로 대체 (이미지가 손실될 수 있습니다)")
                    self.fill_workbook_openpyxl(template_path, context, output_path)
                else:
                    print("잘못된 선택. openpyxl로 대체합니다.")
                    self.fill_workbook_openpyxl(template_path, context, output_path)
            except (EOFError, KeyboardInterrupt):
                print("입력 없음. openpyxl로 대체합니다.")
                self.fill_workbook_openpyxl(template_path, context, output_path)

    def fill_workbook_openpyxl(self, template_path, context, output_path):
        """openpyxl을 사용한 기본 방식 (백업용)"""
        print(f"\n템플릿 처리 (openpyxl): {template_path}")
        
        # 출력 디렉토리 생성
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        
        # 1단계: 템플릿 파일을 출력 위치로 직접 복사
        try:
            shutil.copy2(template_path, output_path)
            print(f"  템플릿 복사 완료: {output_path}")
        except Exception as e:
            print(f"  템플릿 복사 실패: {e}")
            raise
        
        # 2단계: 복사된 파일을 열어서 플레이스홀더만 치환
        try:
            wb = load_workbook(output_path, data_only=False)
            print("  복사된 파일 로드 완료")
            
            # 모든 시트에서 플레이스홀더 치환
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                print(f"  시트 처리: {sheet_name}")
                
                placeholder_count = 0
                
                # 플레이스홀더 치환
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str) and '{{' in cell.value:
                            original_value = cell.value
                            self.replace_placeholders_in_cell(cell, context)
                            placeholder_count += 1
                            if original_value != cell.value:
                                print(f"    치환 {placeholder_count}: {original_value[:50]}... -> {cell.value[:50]}...")
                
                print(f"  {sheet_name} 시트: {placeholder_count}개 플레이스홀더 처리 완료")
            
            # 3단계: Excel 저장
            wb.save(output_path)
            print(f"Excel 저장 완료: {output_path}")
            
            # 4단계: PDF 저장 시도 (xlwings 사용)
            save_pdf_option = self.config.get("save_pdf", True)
            print(f"  PDF 저장 설정: {save_pdf_option}")
            
            if save_pdf_option:
                pdf_path = output_path.replace('.xlsx', '.pdf')
                try:
                    self.save_as_pdf_xlwings(output_path, pdf_path)
                except Exception as e:
                    print(f"PDF 저장 실패 (Excel 필요): {e}")
                    import traceback
                    print(f"상세 오류: {traceback.format_exc()}")
            else:
                print("  PDF 저장 건너뛰기 (설정에서 비활성화)")
            
        except Exception as e:
            print(f"데이터 처리 실패: {e}")
            # 복사된 파일 삭제 (실패 시)
            if os.path.exists(output_path):
                os.remove(output_path)
            raise
    
    # 더 이상 사용하지 않음 - 파일 복사 방식으로 변경
    # def preserve_images(self, worksheet):
    
    # 더 이상 사용하지 않음 - 파일 복사 방식으로 변경  
    # def restore_images(self, worksheet, images_info):
    
    def process_all(self):
        """전체 처리 실행"""
        print("=" * 60)
        print("입사지원서 자동 작성 도구 시작")
        print("=" * 60)
        
        # 파일 경로 확인
        template_path = self.config["template_file"]
        raw_data_path = self.config["raw_data_file"]
        
        if not os.path.exists(template_path):
            print(f"템플릿 파일이 없습니다: {template_path}")
            return
            
        if not os.path.exists(raw_data_path):
            print(f"원천 데이터 파일이 없습니다: {raw_data_path}")
            return
        
        # 원천 데이터 로드
        print(f"\n원천 데이터 로드: {raw_data_path}")
        try:
            df = pd.read_excel(
                raw_data_path, 
                sheet_name=self.config["raw_data_sheet"],
                dtype=str  # 모든 컬럼을 문자열로 읽기
            ).fillna("")
            
            print(f"데이터 로드 완료: {len(df)}행, {len(df.columns)}개 컬럼")
            print(f"컬럼 목록: {list(df.columns)}")
            
        except Exception as e:
            print(f"데이터 로드 실패: {e}")
            return
        
        # 출력 디렉토리 생성
        output_dir = self.config["output_dir"]
        os.makedirs(output_dir, exist_ok=True)
        
        # 각 행별로 지원서 생성
        print(f"\n지원서 생성 시작...")
        success_count = 0
        
        for index, row in df.iterrows():
            try:
                # 컨텍스트 생성 (행 데이터를 딕셔너리로 변환)
                context = row.to_dict()
                
                # 파일명 생성
                try:
                    filename = self.config["filename_pattern"].format(**context)
                except KeyError as e:
                    print(f"  행 {index+1}: 파일명 패턴에 필요한 필드 없음 {e}")
                    filename = f"application_{index+1}.xlsx"
                
                output_path = os.path.join(output_dir, filename)
                
                print(f"\n 행 {index+1}/{len(df)} 처리 중...")
                print(f"  대상: {context.get('이름', 'Unknown')}")
                
                # 템플릿 채우기
                self.fill_workbook(template_path, context, output_path)
                success_count += 1
                
            except Exception as e:
                print(f" 행 {index+1} 처리 실패: {e}")
                continue
        
        print("\n" + "=" * 60)
        print(f"처리 완료! 총 {success_count}/{len(df)}개 파일 생성")
        print(f"출력 폴더: {os.path.abspath(output_dir)}")
        print("=" * 60)
    
    def show_sample_data(self, rows=3):
        """원천 데이터 샘플 출력"""
        raw_data_path = self.config["raw_data_file"]
        
        if not os.path.exists(raw_data_path):
            print(f"원천 데이터 파일이 없습니다: {raw_data_path}")
            return
            
        df = pd.read_excel(raw_data_path, sheet_name=self.config["raw_data_sheet"])
        print(f"\n원천 데이터 샘플 ({raw_data_path}):")
        print("-" * 50)
        print(f"총 {len(df)}행, {len(df.columns)}개 컬럼")
        print(f"컬럼: {list(df.columns)}")
        print(f"\n상위 {rows}행:")
        print(df.head(rows).to_string(index=False))


def main():
    """메인 실행 함수"""
    filler = ExcelTemplateFiller()
    
    # 명령행 인자 처리
    import sys
    
    if len(sys.argv) > 1:
        command = sys.argv[1].lower()
        if command == "sample":
            filler.show_sample_data()
            return
        elif command == "config":
            print("현재 설정:")
            print(json.dumps(filler.config, ensure_ascii=False, indent=2))
            return
    
    # 기본 실행
    filler.process_all()


if __name__ == "__main__":
    main()
